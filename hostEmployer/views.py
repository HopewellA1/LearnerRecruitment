from django.shortcuts import render
from django.contrib import messages
from django.shortcuts import render,redirect, get_object_or_404
from applicants.models import *
from .models import *
from django.contrib.auth.decorators import login_required
from django.contrib.sites.shortcuts import get_current_site
import openpyxl
from openpyxl.styles import PatternFill
from django.contrib.sites.shortcuts import get_current_site
from io import BytesIO
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from rest_framework.decorators import api_view
from rest_framework.response import Response 
from django.core.exceptions import ObjectDoesNotExist




@login_required
def companyDashboard(request):
    
    user = request.user
    try:
        exac = get_object_or_404(Exac, user = user)
    except:
        messages.error(request, "you have not joined a company, please Select or add  your company below")
        return redirect("SelectCompany")
    
    company = exac.Company
    return redirect("departments", CompanyId= company.CompanyId )
        
    
    
@login_required
def SelectCompany(request):
    user  = get_object_or_404(User, username = request.user)
    tour = None
    if user.is_staff == False:
        tour = getPageTour(user, "Path")
       
            
    companies = Company.objects.all()
    if request.method == 'GET':
        
        payload = {
            "companies":companies,
            "tour":tour,
            "domain": get_current_site(request)
        }
        return render(request, 'hostEmployer/SelectCompany.html', payload)
    
    
    if request.method == 'POST':
        try:
            company = get_object_or_404(Company, user  = user)
            profile = {
                
                "company":company,
                "FirstName": request.POST["FirstName"],
                "LastName":  request.POST["LastName"],
                "Gender":request.POST["Gender"],
                "Position": request.POST["Position"],
                "EmailAddress": request.POST["EmailAddress"],
                "PhoneNumber":request.POST["PhoneNumber"],
                "Address": request.POST["Address"]
            }
            saveExacInfo(profile, request)
            msg = f"You have already added a company({company.Name}) you may add divisions for placements."   
        except:        
            company = Company.objects.create(
                user  = user,
                Name = request.POST["Name"],
                Phone = request.POST["Phone"],
                Email = request.POST["Email"]
            ) 
        
            profile = {
                
                "company":company,
                "FirstName": request.POST["FirstName"],
                "LastName":  request.POST["LastName"],
                "Gender":request.POST["Gender"],
                "Position": request.POST["Position"],
                "EmailAddress": request.POST["EmailAddress"],
                "PhoneNumber":request.POST["PhoneNumber"],
                "Address": request.POST["Address"]
            }
            isCreatedProfile = saveExacInfo(profile, request) 
            msg = 'The company details and your profile information have been added successfully. You can now add your company\'s divisions or departments.'
        
        messages.success(request,msg)
        #return redirect('addManagement',CompanyId =company.CompanyId )
        return redirect('departments', CompanyId = company.CompanyId)
    

@api_view(['GET'])
def searchCompany(request, searched):
    companies = []
    searched = searched.strip().lower()
    # Perform the search query on the Company model
    results = Company.objects.filter(
        Q(Name__icontains=searched) 
    )
       
    for company in results:
        print("company::::: ",company)
        companies.append( {
            "CompanyId": company.CompanyId,
            "Name": company.Name,
            "Phone": company.Phone,
            "Email": company.Email,
            "Address": company.Address,
            "dateAdded": company.dateAdded.isoformat(),
            "first_name": company.user.first_name,
            "last_name": company.user.last_name ,
            "email": company.user.email ,
        })
    print("companies: ", companies)    
    return JsonResponse({"companies": companies}, status=200)
    
    
def saveExacInfo(profile, request):
   
    try:
        exac = get_object_or_404(Exac, user = request.user)
    except:
        exac = Exac.objects.create(
            user = request.user,
            Company = profile["company"],
            FirstName =profile["FirstName"],
            LastName =profile["LastName"],
            Gender = profile["Gender"],
            Position =profile["Position"],
            EmailAddress = profile["EmailAddress"],
            PhoneNumber = profile["PhoneNumber"],
            Address = profile["Address"],   
        )
        
    if exac:
        return False
    else:
        return False
        
        
        
@login_required  
def addManagement(request,CompanyId):
    company = get_object_or_404(Company, pk = CompanyId)
    user = request.user
    
    if request.method =='GET':
        
        payload = {
           "company": company,
            "tour":getPageTour(user, "Profile") 
        }
        
        return render(request, 'hostEmployer/addManagement.html', payload)
    
    if request.method == 'POST':
        try:
            exac = get_object_or_404(Exac, user = user)
        except:
            exac = Exac.objects.create(
                user = user,
                Company = company,
                FirstName = request.POST["FirstName"],
                LastName = request.POST["LastName"],
                Gender = request.POST["Gender"],
                Position = request.POST["Position"],
                EmailAddress = request.POST["EmailAddress"],
                PhoneNumber = request.POST["PhoneNumber"],
                Address = request.POST["Address"],   
            )
        
        messages.success(request, 'Manager profile details added successfully, please add or select your Division to add learners.')
        return redirect('departments', CompanyId = CompanyId)
    
    
@login_required   
def departments(request, CompanyId):
    
    company = get_object_or_404(Company, pk = CompanyId)
    Departments = Department.objects.filter(Company = company)
    departments = []
    try:
        exac = get_object_or_404(Exac, user = request.user)
    except:
        exac = None        
    for department in Departments:
        departments.append(
            {
               "department":  department,
               "numLearners": calcList(Learner.objects.filter(Department=department)),
               
            }
        )
        

    
    if request.method == 'GET':
        hasDivisions = False
        if departments:
            hasDivisions = True
        payload = {
            "company": company,
            "departments":departments,
            "exac": exac,
            "tour":getPageTour(request.user,"NewDivision"),
            "basicTour":getPageTour(request.user,"BasicTour"),
            "domain": get_current_site(request),
            "hasDivisions":hasDivisions
           
            
        }
        
        return render(request, 'hostEmployer/departments.html', payload)
    
    if request.method == 'POST':
        
        department = Department.objects.create(
            Company = company,
            Exac= exac,
            Name = request.POST["Name"],
            Description = request.POST["Description"]   
        )
        messages.success(request, "Division added, you may add learners to Division")
        return redirect("departments",CompanyId = company.CompanyId )
    
@login_required   
def DepartmentLearners(request,departmentId):
    try:
        department = get_object_or_404(Department, pk = departmentId)
    except:
        messages.error(request, "The division you are trying to access was not found.")

        return redirect("companyDashboard")
      
    
    company = department.Company
    try:
        exac = get_object_or_404(Exac, user = request.user)
    except:
        exac = None
    learners = Learner.objects.filter(Department = department, Status="Recruited")
    #tour 
    if request.method == 'GET':
        
        payload = {
           "RecruitedLearners": learners,
            "department": department,
            "company":company,
            "exac": exac,
            "recruited":True,
            "tour":getPageTour(request.user,"NewDivision"),
        }
        
        return render(request, 'applicants/learners.html', payload)
    
    
@login_required
def selectCategory(request, departmentId):
    user = request.user
    try:
        department = get_object_or_404(Department, pk = departmentId)  
    except:
        messages.error(request, 'The division you tried to access was not found')
        return redirect('companyDashboard')  
    Categories = category.objects.all()
    print("Categories: ", Categories)
    payload = {
        "department": department,
        "categories": Categories
    }
    
    messages.warning(request, f"Please select a category from the options below to add learners to your division:  {department.Name}")
    return render(request, 'hostEmployer/selectCategory.html', payload)

@login_required
def selectLearner(request, departmentId, categoryId):
    try:
        department = get_object_or_404(Department, pk = departmentId)  
    except:
        messages.error(request, 'The division you tried to access was not found')
        return redirect('companyDashboard') 
    Category = get_object_or_404(category, pk= categoryId)
    learners = Learner.objects.filter(category = Category)
    availableLearners = Learner.objects.filter(category = Category, Status ="Available")
    RecruitedLearners = Learner.objects.filter(category = Category, Status ="Recruited")
    if request.method == 'GET':
        
        payload = {
           "learners": learners,
            "Category": Category,
            "department": department,
            "selectState":True,
            "AvailableLearners": availableLearners,
            "RecruitedLearners": RecruitedLearners
        }

        return render(request, 'applicants/learners.html', payload)

@login_required
def ConfirmLearner(request, learnerId, departmentId):
    try:
        department = get_object_or_404(Department, pk = departmentId)  
    except:
        messages.error(request, 'The division you tried to access was not found')
        return redirect('companyDashboard')
    
    try:
        learner = get_object_or_404(Learner, pk=learnerId)
    except:
        messages.error(request, 'The learner you tried to access was not found')
        return redirect('companyDashboard')
    Category = learner.category
    try:
        exac = get_object_or_404(Exac, user = request.user)
    except:
        messages.error(request, "You must be a manager to take this action.")
        return redirect("home")

    if request.method == 'GET':
        
        payload = {
            "learner": learner,
            "Category":Category,
            "department":department,
            "selectState":True
        }
        messages.warning(request, f"Are you sure you want to add this learner to your division({department.Name})?")
        return render(request, 'applicants/learnerDetails.html',payload )
    
    if request.method == 'POST':
        
        learner.Status = "Recruited"
        learner.Exac = exac
        learner.Department = department
        learner.save()
        
        messages.success(request, "The Learner has been added to the division")
        return redirect("selectLearner", departmentId = department.DepartmentId , categoryId =learner.category.categoryId )
@login_required      
def confirmRemoval(request, learnerId):
    
    try: 
        learner =get_object_or_404(Learner, pk =learnerId )
    except:
        messages.error(request, "The learner you tried to access was not found")
        return redirect("companyDashboard")
    
    department = learner.Department
    Category = learner.category
    
    if request.method == 'GET':
        payload = {
            "learner": learner,
            "Category":Category,
            "department":department,
            "is_remove":True
        }
        
        messages.warning(request, f"Are you sure you want to remove this learner from the division of {department.Name} at {department.Company.Name}")
        return render(request, 'applicants/learnerDetails.html',payload)

    
    if request.method == 'POST':
    #removing learner
        learner.Department = None
        learner.Exac = None
        learner.Status = "Available"
        learner.save()
        messages.success(request, "The learner has been removed from the division")
        return redirect("DepartmentLearners", departmentId = department.DepartmentId )
    
def calcList(list):
    numObj = 0
    for obj in list:
        numObj +=1
    return numObj



#compnay CRUD

@login_required
def companyDetails(request, companyId):
    
    try:
        company = get_object_or_404(Company, pk = companyId)
    except:
        messages.error(request, "The record you tried to access was not found")
        return redirect("home")
    
    try:
        exac = get_object_or_404(Exac, user = request.user,Company=company)
    except:
        exac = None    

    All_Exacs = Exac.objects.filter(Company = company)
    
    
    if request.method == 'GET':
        
        payload = {
            "company":company,
            "exac":exac  ,
            "All_Exacs":All_Exacs,
            "numManagers":calcList(All_Exacs) ,
            "numLearners":  calcCompanyLeanrners(company.CompanyId)
        }
        return render(request, 'hostEmployer/companyDetails.html', payload)
        

def calcCompanyLeanrners(companyId):
    nunLearners = 0
    for dep in Department.objects.filter(Company = get_object_or_404(Company, pk = companyId)):
        nunLearners += calcList(Learner.objects.filter(Department=dep))  
    return nunLearners
    
    
@login_required       
def EditCompany(request, companyId):
    try:
        company = get_object_or_404(Company, pk = companyId)
    except:
        messages.error(request, "The record you tried to access was not found")
        return redirect("home")
    
    try:
        exac = get_object_or_404(Exac, user = request.user,Company=company)
    except:
        exac = None    

    
    
    if request.method == 'GET':
        
        payload = {
            "company":company,
            "exac":exac   
        }
        return render(request, 'hostEmployer/EditCompany.html', payload)
    
    
    if request.method == 'POST':
        
        company.Name = request.POST["Name"]
        company.Phone = request.POST["Phone"]
        company.Email = request.POST["Email"]
        company.Address = request.POST["Address"]
        
        company.save()
        messages.success(request, "Changes saved successfully.")
        return redirect("companyDetails", companyId=company.CompanyId)
         
    
@login_required   
def departmentDetails(request, departmentId):
    
    try:
        department = get_object_or_404(Department, pk = departmentId)
        
    except:
        messages.error(request, "The division you tried to access was not found.")
        return redirect("home")
   
    try:
        exac = get_object_or_404(Exac, user = request.user,Company=company)
    except:
        exac = None
    
    company = department.Company
    otherDepartments =[]
    Departments = Department.objects.filter(Company = company)
    for department1 in Departments:
        otherDepartments.append(
            {
               "department":  department1,
               "numLearners": calcList(Learner.objects.filter(Department=department1)),
               
            }
        )
    if request.method == 'GET':
      
        payload = {
            "department": department,
            "company": company,
            "exac":exac,
            "otherDepartments":otherDepartments
            
        }
        return render(request, 'hostEmployer/departmentDetails.html', payload)
    
    if request.method == 'POST':
        
        department.Name = request.POST["Name"]
        department.Description = request.POST["Description"]
        department.save()
        messages.success(request, "Changes have been saved.")
        return redirect("departmentDetails", departmentId=department.DepartmentId)
    

@login_required         
def CompanyLearners(request, CompanyId):
    
    try:
        company = get_object_or_404(Company, pk = CompanyId)
    except:
        messages.error(request, "The records you tried to access were not found.") 
        return redirect("home")
    payload = {
        "learners": getCompanyLearners(company.CompanyId),
        "company":company
    }
    return render(request, 'hostEmployer/CompanyLearners.html', payload)
    
      
def getCompanyLearners(CompanyId):
    company = get_object_or_404(Company, pk = CompanyId) 
    learners = []
    deps = Department.objects.filter(Company= company)
    for dep in deps:
        for learner in Learner.objects.filter(Department = dep):
            learners.append(learner)
    return learners


def download_departmentExcel(request):
    departmentId = request.GET.get('departmentId')
    companyId = request.GET.get('companyId')
    spreadName = ""
    if companyId:
        Learners = []
        
        
        try:
            company = get_object_or_404(Company, pk = companyId)
        except:
            messages.error(request, "The company was not found.")
            return redirect("home")
        
        departments = Department.objects.filter(Company = company)
        for dep in departments:
            
            for i in Learner.objects.filter(Department = dep):
               Learners.append(i) 
               
        spreadName = company.Name+"_learners"
        returning  = "comp"
        
          
    
    if departmentId:


      
        try:
            department = get_object_or_404(Department, pk = departmentId)
        except:
            messages.error(request, "The division was not found.")
            return redirect("home")
        Learners = Learner.objects.filter(Department = department)
        
        spreadName = department.Company.Name+'_'+department.Name+'_learners'
        
        returning  = "dep"
        
        
    response = None
    if Learners:   
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Define the headers and dummy data
        headers = [ 
                    'Learner ID Number',
                    'Date of birth','Gender',
                    'Equity Code',
                    'Nationality Code',
                    'Home Language',
                    'Learner Surname',
                    'Learner First Name',
                    'Learner Middle Name',
                    'RACE',
                    'Learner Home Address1',
                    'Learner Home Address2',
                
                    'Learner Home Postal Code',
                    'STATSSA Area',
                    'Learner Cell PhoneNumber',
                    'LearnerEmailAddress',
                    'Learner Fax Number',
                    'Province',
                    'Disability',
                    'LastSchool EMIS No',
                    'Last School Name',
                    'Last School Year', 
                    'Degree Title',
                    'Institution',
                    'Field Of Study',
                    'NQF Level',
                    'Experience',
                    'Status',
                    'Company',
                    'Division'
        ]
        data = []
        for learner in Learners:
            data.append(
                [ 
                    learner.LearnerIDNumber,
                    learner.DOB,
                    learner.Gender,
                    learner.EquityCode,
                    learner.NationalityCode,
                    learner.HomeLanguage,
                    learner.LearnerSurname,
                    learner.LearnerFirstName,
                    learner.LearnerMiddleName,
                    learner.RACE,
                    learner.LearnerHomeAddress1,
                    learner.LearnerHomeAddress2,
                    learner.LearnerHomePostalCode,
                    learner.STATSSAArea,
                    learner.LearnerCellPhoneNumber,
                    learner.LearnerFaxNumber,
                    learner.LearnerEmailAddress,
                    learner.Province,
                    learner.Disability,
                    learner.LastSchoolEMISNo,
                    learner.LastSchoolName,
                    learner.LastSchoolYear,
                    learner.DegreeTitle,
                    learner.Institution,
                    learner.FieldOfStudy,
                    learner.NQFLevel,
                    learner.Experience,
                    learner.Status,
                    learner.Department.Company.Name,
                    learner.Department.Name,
                        
                ]
            )
        

        # Populate the header row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            # Style the header cell with a green background
            cell.fill = PatternFill(start_color="60FF5C", end_color="60FF5C", fill_type="solid")

        # Populate the data rows
        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)

        # Save the workbook to a BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Set up the response
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename='+spreadName+'.xlsx'
        print("here")
        
    if response:
        return response
    else:
        if returning == "comp":
            messages.warning(request, "The company has not recruited any learners.")
            return redirect("CompanyLearners", CompanyId =companyId )
        elif returning == "dep":
            messages.warning(request, "The department has not recruited any learners.")
            
            return redirect("DepartmentLearners", departmentId =departmentId )

    
        
        
        
@login_required 
def SearchLearners(request):
    try:
        Searched = request.POST["Searched"].strip()
    except:
       Searched = ''
    if request.method == "POST":
        Learners = []
        if request.POST["space"] == "Company":
            companyId = int(request.POST["companyId"])
            try:
                company = get_object_or_404(Company, pk = companyId)
            except:
                messages.error(request, "The query company was not found.")
                return redirect("home")
            
            departments = Department.objects.filter(Company=company)
            numDeps = 0
            for dep in departments:
                numDeps += 1
                l = Learner.objects.filter(
                   Q(Department=dep) & 
                   (
                        Q(LearnerFirstName__icontains=Searched) |
                        Q(LearnerSurname__icontains=Searched) |
                        Q(NQFLevel__icontains=Searched) |
                        Q(Municipality__icontains=Searched) |
                        Q(STATSSAArea__icontains=Searched) |
                        Q(Institution__icontains=Searched) |
                        Q(Status__icontains=Searched) |
                        Q(MajorSubjects__icontains=Searched) |
                        Q(Experience__icontains=Searched) |
                        Q(LearnerHomeAddress1__icontains=Searched) |
                        Q(LearnerIDNumber__icontains=Searched) |
                        Q(DegreeTitle__icontains=Searched) |
                        Q(FieldOfStudy__icontains=Searched)
                    )
                )
                for i in l:
                    Learners.append(i)    
                   
            numLearners = calcList(Learners)
        
            if numLearners> 0:
                if numLearners == 1:
                    messages.success(request, f"{numLearners} learner found on search: {Searched}")
                else:
                    messages.success(request, f"{numLearners} learners found on search: {Searched}")
            else:
                messages.error(request, f"No learners found on search: {Searched}")  
            payload = {
                "learners": Learners,
                "company":company,
                "isSearch": True
            }
            return render(request, 'hostEmployer/CompanyLearners.html', payload)       
        elif request.POST["space"] == "Department":
            try:
                department = get_object_or_404(Department,pk = int(request.POST["departmentId"]))
            except:
                messages.error(request, "The query Division was not found.")
                return redirect("home")    
            
            
            Learners = Learner.objects.filter(
                Q(Department = department) &  (
                        Q(LearnerFirstName__icontains=Searched) |
                        Q(LearnerSurname__icontains=Searched) |
                        Q(NQFLevel__icontains=Searched) |
                        Q(Municipality__icontains=Searched) |
                        Q(STATSSAArea__icontains=Searched) |
                        Q(Institution__icontains=Searched) |
                        Q(Status__icontains=Searched) |
                        Q(MajorSubjects__icontains=Searched) |
                        Q(Experience__icontains=Searched) |
                        Q(LearnerHomeAddress1__icontains=Searched) |
                        Q(LearnerIDNumber__icontains=Searched) |
                        Q(DegreeTitle__icontains=Searched) |
                        Q(FieldOfStudy__icontains=Searched)
                    )
            )
            
            numLearners = calcList(Learners)
            
            if numLearners> 0:
                if numLearners == 1:
                    messages.success(request, f"{numLearners} learner found on search: {Searched}")
                else:
                    messages.success(request, f"{numLearners} learners found on search: {Searched}")
            else:
                messages.error(request, f"No learners found on search: {Searched}") 
            
            print("Learners: ", Learners)
            try:
                exac = get_object_or_404(Exac, user = request.user)
            except:
                exac = None
            
            payload = {
                "learners": Learners,
                "department": department,
                "company":department.Company,
                "exac": exac,
                "isSearch": True
            }
            
            return render(request, 'applicants/learners.html', payload)
        elif request.POST["space"] == "Pool":
            try:
                Category = get_object_or_404(category, pk = int(request.POST["CategoryId"]))
            except:
                messages.error(request, "The query category was not found.")
                return redirect("home")    
            Learners = Learner.objects.filter(
                Q(category = Category) &  (
                        Q(LearnerFirstName__icontains=Searched) |
                        Q(LearnerSurname__icontains=Searched) |
                        Q(NQFLevel__icontains=Searched) |
                        Q(Municipality__icontains=Searched) |
                        Q(STATSSAArea__icontains=Searched) |
                        Q(Institution__icontains=Searched) |
                        Q(Status__icontains=Searched) |
                        Q(MajorSubjects__icontains=Searched) |
                        Q(Experience__icontains=Searched) |
                        Q(LearnerHomeAddress1__icontains=Searched) |
                        Q(LearnerIDNumber__icontains=Searched) |
                        Q(DegreeTitle__icontains=Searched) |
                        Q(FieldOfStudy__icontains=Searched)
                    )
            )
            numLearners = calcList(Learners)
            
            if numLearners> 0:
                if numLearners == 1:
                    messages.success(request, f"{numLearners} learner found on search: {Searched}")
                else:
                    messages.success(request, f"{numLearners} learners found on search: {Searched}.")
            else:
                messages.error(request, f"No learners found on search: {Searched}") 
            payload = {
                "learners": Learners,
                "Category": Category,
                "isSearch": True
            }
            return render(request, 'applicants/learners.html', payload)
        elif request.POST["space"] == "select":
            
            try:
                Category = get_object_or_404(category, pk = int(request.POST["CategoryId"]))
            except:
                messages.error(request, "The query category was not found.")
                return redirect("home")
            try:
                department = get_object_or_404(Department, pk = int(request.POST["departmentId"]))
            except:
                messages.error(request, "The query category was not found.")
                return redirect("home")    
            Learners = Learner.objects.filter(
                Q(category = Category) &  (
                        Q(LearnerFirstName__icontains=Searched) |
                        Q(LearnerSurname__icontains=Searched) |
                        Q(NQFLevel__icontains=Searched) |
                        Q(Municipality__icontains=Searched) |
                        Q(STATSSAArea__icontains=Searched) |
                        Q(Institution__icontains=Searched) |
                        Q(Status__icontains=Searched) |
                        Q(MajorSubjects__icontains=Searched) |
                        Q(Experience__icontains=Searched) |
                        Q(LearnerHomeAddress1__icontains=Searched) |
                        Q(LearnerIDNumber__icontains=Searched) |
                        Q(DegreeTitle__icontains=Searched) |
                        Q(FieldOfStudy__icontains=Searched)
                    )
            )
            numLearners = calcList(Learners)
            
            if numLearners> 0:
                if numLearners == 1:
                    messages.success(request, f"{numLearners} learner found on search: {Searched}")
                else:
                    messages.success(request, f"{numLearners} learners found on search: {Searched}.")
            else:
                messages.error(request, f"No learners found on search: {Searched}")
                
            payload = {
            "learners": Learners,
                "Category": Category,
                "department": department,
                "selectState":True,
                "isSearch": True
                
            }

            return render(request, 'applicants/learners.html', payload)
   
@api_view(['POST', 'GET'])
def takenTour(request, tourId):
    
    if request.method =="GET":
        try:
            
            tour = get_object_or_404(Tour, pk= tourId)
            tour.is_taken =True
            tour.save()
            status = "success"
        except:
            status = "fail"
            
        
        payload = {
            "Status": status,
        }
        return Response(payload)


def getPageTour(user, step):
    #Profile, Path
    
    try:
        tour = get_object_or_404(Tour, user = user, Step = step)
    except: 
        tour = Tour.objects.create(
            user = user,
            Step = step 
        )
    return tour
    