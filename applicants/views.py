from django.shortcuts import render
from django.shortcuts import render,redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import PatternFill
from io import BytesIO
from .models import *
from django.contrib import messages
import pandas as pd
from django.core.files.storage import FileSystemStorage
from django.db.models import Q
from docx import Document
from docx.shared import Inches
from docx.enum.section import WD_ORIENT
from docx.enum.text import WD_ALIGN_PARAGRAPH
from django.conf import settings
#new to fix db
from django.db import transaction
from datetime import datetime



def searchLeaners(request, searchLocation):
    
    pass
    
    

def home(request):
    removeFakeAccounts()
    # with transaction.atomic():
    #     # Group learners by LearnerIDNumber, keeping only the earliest record
    #     unique_learners = (
    #         Learner.objects.values('LearnerIDNumber')
    #         .annotate(min_id=models.Min('ApplicantId'))
    #     )

    #     # Extract unique IDs to keep
    #     unique_ids = [item['min_id'] for item in unique_learners]

    #     # Delete duplicate records, keeping only unique ones
    #     Learner.objects.exclude(ApplicantId__in=unique_ids).delete()

    # print("Duplicates deleted successfully.")

   
    return render(request, 'applicants/home.html')




def LPO5(request):
   # Category = get_object_or_404(category, pk = categoryId)
    if request.method == 'GET':
        
        learners = Learner.objects.filter
        
        payload = {
            "learners": learners
        }
        
        return render(request, 'applicants/LPO5.html', payload)
    
    return render(request, 'applicants/LPO1.html')

def categories(request):
    
    
    if request.method =='GET':
        
        Categories = category.objects.all()
        payload = {
            "Categories": Categories
        } 
        
        return render(request, 'applicants/categories.html', payload)
    elif request.method == 'POST':
        
        Category =  category.objects.create(
            categoryName = request.POST["categoryName"]
        )   
        messages.success(request, 'Category added')
        return redirect('categories')    


def learnercategories(request ):
    
    if request.method == 'GET':
        categories = category.objects.all()
        
        payload = {
           
            "categories":categories
        }
        return render(request, 'applicants/learnercategories.html', payload)
    
        
@login_required   
def Learners(request, categoryId):
    Category = get_object_or_404(category, pk= categoryId)
    
    learners = Learner.objects.filter(category = Category) 
    AvailableLearners = Learner.objects.filter(category = Category, Status="Available") 
    if request.method == 'GET':
        
        payload = {
           "learners": learners,
            "Category": Category,
            "AvailableLearners": AvailableLearners,
            "RecruitedLearners": Learner.objects.filter(category = Category, Status="Recruited")
        }

        return render(request, 'applicants/learners.html', payload)
    
    if request.method == 'POST':
        
        learner= Learner.objects.create(
            user = request.user,
            category = Category,
            LearnerIDNumber = request.POST["LearnerIDNumber"],
            DOB = request.POST["DOB"],
            Gender = request.POST["Gender"],
            EquityCode = request.POST["EquityCode"],
            HomeLanguage = request.POST["HomeLanguage"],
            LearnerSurname = request.POST["LearnerSurname"],
            LearnerFirstName = request.POST["LearnerFirstName"],
            LearnerMiddleName = request.POST["LearnerMiddleName"],
            LearnerHomeAddress1 = request.POST["LearnerHomeAddress1"],
            LearnerHomeAddress2 = request.POST["LearnerHomeAddress2"],
            Municipality=  request.POST["Municipality"],
            LearnerHomePostalCode = request.POST["LearnerHomePostalCode"],
            STATSSAArea = request.POST["STATSSAArea"],
            LearnerCellPhoneNumber = request.POST["LearnerCellPhoneNumber"],
            LearnerFaxNumber = request.POST["LearnerFaxNumber"],
            LearnerEmailAddress = request.POST["LearnerEmailAddress"],
            Province = request.POST["Province"],
            Disability = request.POST["Disability"],
            LastSchoolEMISNo = request.POST["LastSchoolEMISNo"],
            LastSchoolName = request.POST["LastSchoolName"],
            DegreeTitle = request.POST["DegreeTitle"],
            FieldOfStudy = request.POST["FieldOfStudy"],
            NQFLevel = request.POST["NQFLevel"],
            RACE = request.POST["RACE"],
            Institution = request.POST["Institution"],
            LastSchoolYear = request.POST["LastSchoolYear"],
            MajorSubjects = request.POST["MajorSubjects"],
            Experience = request.POST["Experience"],
            
        )
        messages.success(request, "Learner added successfully")
        
        return redirect("Learners", categoryId = Category.categoryId)


def editLearner(request, learnerId):
    
    try:
        learner = get_object_or_404(Learner, pk = learnerId)
    except:
        messages.error(request, "The Learner was not found")
        
    if request.method == 'GET':
        payload = {
            "learner":learner,
            "Category": learner.category
        }
       
        return render(request, 'applicants/editLearner.html', payload)
    
    if request.method =='POST':
        
        numUpdate = 0
        
        if learner.LearnerIDNumber != request.POST["LearnerIDNumber"]:
            learner.LearnerIDNumber = request.POST["LearnerIDNumber"]
            numUpdate += 1
            
        if learner.DOB != request.POST["DOB"]:
            learner.DOB =request.POST["DOB"]
            numUpdate +=1 
            
        if learner.Gender != request.POST["Gender"]:
            learner.Gender = request.POST["Gender"]
            numUpdate += 1
        
        if learner.EquityCode != request.POST["EquityCode"]:
            learner.EquityCode = request.POST["EquityCode"]
            numUpdate +=1 
      
        if learner.HomeLanguage != request.POST["HomeLanguage"]:
            learner.HomeLanguage = request.POST["HomeLanguage"]
            numUpdate += 1
            
        if learner.LearnerSurname != request.POST["LearnerSurname"]:
            learner.LearnerSurname = request.POST["LearnerSurname"]
            numUpdate += 1
        if learner.LearnerFirstName != request.POST["LearnerFirstName"]:
            learner.LearnerFirstName = request.POST["LearnerFirstName"]
            numUpdate += 1
            
        if learner.LearnerMiddleName != request.POST["LearnerMiddleName"]:
            learner.LearnerMiddleName = request.POST["LearnerMiddleName"]
            numUpdate += 1
            
        if learner.LearnerHomeAddress1 != request.POST["LearnerHomeAddress1"]:
            learner.LearnerHomeAddress1 = request.POST["LearnerHomeAddress1"]
            numUpdate += 1
        if learner.LearnerHomeAddress2 != request.POST["LearnerHomeAddress2"]:
            learner.LearnerHomeAddress2 = request.POST["LearnerHomeAddress2"]
            numUpdate += 1
  
        if learner.Municipality != request.POST["Municipality"]:
            learner.Municipality = request.POST["Municipality"]
            numUpdate += 1
            
        if learner.LearnerHomePostalCode != request.POST["LearnerHomePostalCode"]:
            learner.LearnerHomePostalCode = request.POST["LearnerHomePostalCode"]
            numUpdate += 1
            
        if learner.STATSSAArea != request.POST["STATSSAArea"]:
            learner.STATSSAArea = request.POST["STATSSAArea"]
            numUpdate += 1
            
        if learner.LearnerCellPhoneNumber != request.POST["LearnerCellPhoneNumber"]:
            learner.LearnerCellPhoneNumber = request.POST["LearnerCellPhoneNumber"]
            numUpdate += 1
            
        if learner.LearnerFaxNumber != request.POST["LearnerFaxNumber"]:
            learner.LearnerFaxNumber = request.POST["LearnerFaxNumber"]
            numUpdate += 1
            
        if learner.LearnerEmailAddress != request.POST["LearnerEmailAddress"]:
            learner.LearnerEmailAddress = request.POST["LearnerEmailAddress"]
            numUpdate += 1
            
        if learner.DegreeTitle != request.POST["DegreeTitle"]:
            learner.DegreeTitle = request.POST["DegreeTitle"]
            numUpdate += 1

        if learner.FieldOfStudy != request.POST["FieldOfStudy"]:
            learner.FieldOfStudy = request.POST["FieldOfStudy"]
            numUpdate += 1
            
            
        if learner.NQFLevel != request.POST["NQFLevel"]:
            learner.NQFLevel = request.POST["NQFLevel"]
            numUpdate += 1
            
        if learner.RACE != request.POST["RACE"]:
            learner.RACE = request.POST["RACE"]
            numUpdate += 1
            
        if learner.Institution != request.POST["Institution"]:
            learner.Institution = request.POST["Institution"]
            numUpdate += 1
            
        if learner.MajorSubjects != request.POST["MajorSubjects"]:
            learner.MajorSubjects = request.POST["MajorSubjects"]
            numUpdate += 1
            
        if learner.Experience != request.POST["Experience"]:
            learner.Experience = request.POST["Experience"]
            numUpdate += 1
            
        if numUpdate > 0:
            learner.save()
            messages.success(request, 'changes saved')

        else:
            messages.warning(request, 'No changes made')
        return redirect("learnerDetails", learnerId = learner.ApplicantId)
            
def learnerDetails(request, learnerId):
    
    try:
        
        learner = get_object_or_404(Learner, pk = learnerId)
    except:
        
        messages.error(request, 'The Learner was not found.')
        return redirect("home")
    
    if request.method == 'GET':
        
        payload = {
            "learner": learner,
            "Category":learner.category
        }
        return render(request, 'applicants/learnerDetails.html',payload )
    
    
def deleteLearner(request, learnerId):
    
    try:
        
        learner = get_object_or_404(Learner, pk = learnerId)
    except:
        
        messages.error(request, 'The Learner was not found.')
        return redirect("home")
    Category = learner.category
    if request.method == 'GET':
        
        payload = {
            "learner": learner,
            "Category":Category
        }
        return render(request, 'applicants/deleteLearner.html',payload )
    
    if request.method == 'POST':
        
        if int(request.POST["learnerId"]) == learner.ApplicantId:
            learner.delete()
            messages.success(request, 'The learner has been deleted')
           
        else:
            messages.error(request, 'Action not permitted')
            
        return redirect('Learners',categoryId=Category.categoryId )
        
    

def download_excel(request, categoryId):
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    Category = get_object_or_404(category, pk = categoryId)
    Learners = Learner.objects.filter(category = Category)
    # Define the headers and dummy data
    headers = [ 
                    'Learner ID Number',
                    'Date of birth','Gender',
                    'Equity Code',
                    'Nationality Code',
                    'Home Language',
                    'Learner Surname',
                    'Learner First Name',
                    'Learner Middle Name',
                    'RACE',
                    'Learner Home Address1',
                    'Learner Home Address2',
                
                    'Learner Home Postal Code',
                    'STATSSA Area',
                    'Learner Cell PhoneNumber',
                    'Learner Email Address',
                    'Learner Fax Number',
                    'Province',
                    'Disability',
                    'Last School EMIS No',
                    'Last School Name',
                    'Last School Year', 
                    'Degree Title',
                    'Institution',
                    'Field Of Study',
                    'NQF Level',
                    'Experience',
                    'Status',
                    'Company',
                    'Division'
        ]
    data = []
    for learner in Learners:
        
        if learner.Department:
            DepartmentName = learner.Department.Name
            CompanyName = learner.Department.Company.Name
            
        else:
            DepartmentName = "N/A"
            CompanyName = "N/A"
        
        data.append(
            [ 
                    learner.LearnerIDNumber,
                    learner.DOB,
                    learner.Gender,
                    learner.EquityCode,
                    learner.NationalityCode,
                    learner.HomeLanguage,
                    learner.LearnerSurname,
                    learner.LearnerFirstName,
                    learner.LearnerMiddleName,
                    learner.RACE,
                    learner.LearnerHomeAddress1,
                    learner.LearnerHomeAddress2,
                    learner.LearnerHomePostalCode,
                    learner.STATSSAArea,
                    learner.LearnerCellPhoneNumber,
                   
                    learner.LearnerEmailAddress,
                    learner.LearnerFaxNumber,
                    learner.Province,
                    learner.Disability,
                    learner.LastSchoolEMISNo,
                    learner.LastSchoolName,
                    learner.LastSchoolYear,
                    learner.DegreeTitle,
                    learner.Institution,
                    learner.FieldOfStudy,
                    learner.NQFLevel,
                    learner.Experience,
                    learner.Status,
                    CompanyName,
                    DepartmentName,
                        
                ]
            )
    
    # data = [
    #     ['John', 'A.', 'Doe', 'Male', '123-456-7890', 'john.doe@example.com'],
    #     ['Jane', 'B.', 'Smith', 'Female', '234-567-8901', 'jane.smith@example.com'],
    #     ['Jim', 'C.', 'Brown', 'Male', '345-678-9012', 'jim.brown@example.com'],
    #     ['Jill', 'D.', 'Davis', 'Female', '456-789-0123', 'jill.davis@example.com']
    # ]

    # Populate the header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        # Style the header cell with a green background
        cell.fill = PatternFill(start_color="60FF5C", end_color="60FF5C", fill_type="solid")

    # Populate the data rows
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Set up the response
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename='+Category.categoryName+'.xlsx'
    
    
    if response:
        
        return response
    else:
        messages.warning(request,"The category does not have any learners loaded.")
        return redirect("Learners", categoryId)


def save_excel_to_db(request):
    user = request.user
    try:
        Category = get_object_or_404(category, pk = int(request.POST["categoryId"]))
    except:
        messages.error(request, "The category you tried to load learners to was not found.")
        return redirect("home")
    # Read the Excel file from the file-like object
    df = pd.read_excel(request.FILES["excelData"])
    df.columns = df.columns.str.strip()
 
    numLearners = 0
    numThere = 0
    # Iterate over the DataFrame rows
    for index, row in df.iterrows():
       
        try:
            is_learneron  = get_object_or_404(Learner,LearnerIDNumber = "".join([char for char in str(row['IDENTITY \nNUMBER']) if char.isdigit()]) )
            if is_learneron.LearnerCellPhoneNumber == "nan / nan":
                is_learneron.LearnerCellPhoneNumber = "".join([char for char in str(row['CONTACT \nNUMBER']) if char.isdigit()]) 
                print("number updated: ".join([char for char in str(row['CONTACT \nNUMBER']) if char.isdigit()]))
                is_learneron.save()
            numThere += 1
            if is_learneron.DOB == None:
                is_learneron.DOB= get_dob_from_sa_id(is_learneron.LearnerIDNumber)
                print("is_learneron.DOB: ", is_learneron.DOB)
                is_learneron.save()
        except:
            learner = Learner.objects.create(
                user = user,
                category = Category,
                Gender =  row['GENDER'],
                LearnerIDNumber = "".join([char for char in str(row['IDENTITY \nNUMBER']) if char.isdigit()]) ,
                HomeLanguage = "Zulu",
                LearnerSurname = row['SURNAME'],
                
                LearnerFirstName =  row['NAME'],
                LearnerMiddleName = row['ALTENATIVE Name'],
                LearnerHomeAddress1 = row['PHYSICAL\nADDRESS'],
                Municipality = row['MUNICIPALITY'],
                LearnerHomePostalCode = "N/A",
                STATSSAArea = "N/A",
                LearnerCellPhoneNumber = "".join([char for char in str(row['CONTACT \nNUMBER']) if char.isdigit()]) ,
                LearnerFaxNumber = "N/A",
                LearnerEmailAddress = row['EMAIL \nADDRESS'],
                Province = "N/A",
                Disability = "None",
                LastSchoolEMISNo = "N/A",
                LastSchoolName = "N/A",
                #LastSchoolYear = "N/A",
                DegreeTitle = row['Degree Title'],
                FieldOfStudy =  row['Field of Study'],
                NQFLevel = row['NQF Level'],
                Institution = row['Institution'],
                RACE = row['RACE'],
                MajorSubjects = "N/A",
                Experience = row['Experience'],
                
            )
            numLearners += 1

       
        pass
    messages.success(request, f"{numLearners} learner(s) have been added to the database.")
    if numThere > 0:
        messages.warning(request, f"{numThere} learner(s) have been already added.")
    return redirect("Learners", categoryId =Category.categoryId)
  
  
  

def get_dob_from_sa_id(id_number):
  
    # Extract the YYMMDD part for DOB
    dob_str = id_number[:6]
    # Determine the century and parse the DOB string to a datetime object
    try:
        # Attempt to parse as 19XX (1900s)
        dob = datetime.strptime("19" + dob_str, "%Y%m%d")
    except ValueError:
        # If the above fails, assume 20XX (2000s)
        dob = datetime.strptime("20" + dob_str, "%Y%m%d")
    return dob
  
def filterDigits(string):
    
    pass


def help(request):
    
    return render(request, 'applicants/help.html')



def removeFakeAccounts():
    users = User.objects.filter(is_active = False)
    for user in users:
        pass#print(f"Name: {user.first_name}; last name: {user.last_name};**email: {user.email} ;**** Active State: {user.is_active}; super user state: {user.is_superuser}")
        #user.delete()